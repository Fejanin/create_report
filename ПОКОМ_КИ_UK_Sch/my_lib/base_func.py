import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side
from my_lib.table import Table, Row
from openpyxl.utils import get_column_letter


style_name = 'Arial10px'
new_style = NamedStyle(name=style_name)
new_style.font = Font(name='Arial', size=10)
RED_FONT = 'ff0000'
RED_CELL = 'f05959'
NUM_FORMAT = '0_ ;[Red]\-0\ ' # настройка ячеек: число без десятых, отрицательные красным цветом
HEADER_COLOR = 'FFF4C5'
HEADER_COLOR_SUM = '758ce0'

def read_xlsx_file(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row
    return wb, ws, max_col, max_row

def create_base_table_from_dv_file(file_dv):
    table = Table()
    wb, ws, max_col, max_row = read_xlsx_file(file_dv)

    all_columns = {
        'name_sku': None,  # Номенклатура
        'units_of_measurement': None,  # Ед. изм.
        'initial_balance': None,  # Начальный остаток
        'receipt_of_products': None,  # Приход
        'final_balance': None,  # Конечный остаток
    }

    flag = False
    for row in range(1, max_row + 1):
        current_name_row = None
        for col in range(1, max_col + 1):
            if all(all_columns.values()):
                flag = True
            cell_value = ws.cell(row=row, column=col).value
            if cell_value == 'Итого':
                flag = False
                break
            if not all(all_columns.values()):
                if cell_value == 'Номенклатура':
                    all_columns['name_sku'] = all_columns['name_sku'] or col
                if cell_value == 'Ед. изм.':
                    all_columns['units_of_measurement'] = all_columns['units_of_measurement'] or col
                if cell_value == 'Начальный остаток':
                    all_columns['initial_balance'] = all_columns['initial_balance'] or col
                if cell_value == 'Приход':
                    all_columns['receipt_of_products'] = all_columns['receipt_of_products'] or col
                if cell_value == 'Конечный остаток':
                    all_columns['final_balance'] = all_columns['final_balance'] or col
            if flag and col in all_columns.values():
                if col == all_columns['name_sku']:
                    current_name_row = cell_value
                    r = Row(current_name_row)
                    if r:
                        table.add_row(r)
                elif current_name_row and col == all_columns['units_of_measurement']:
                    table.rows[table.rows.index(current_name_row)].units_of_measurement = cell_value
                elif current_name_row and col == all_columns['initial_balance']:
                    table.rows[table.rows.index(current_name_row)].initial_balance = cell_value
                elif current_name_row and col == all_columns['receipt_of_products']:
                    table.rows[table.rows.index(current_name_row)].receipt_of_products = cell_value
                elif current_name_row and col == all_columns['final_balance']:
                    table.rows[table.rows.index(current_name_row)].final_balance = cell_value
    return table


# 'f05959'
def add_data_from_old_file(file_dv_old, table):
    wb, ws, max_col, max_row = read_xlsx_file(file_dv_old)

    all_columns = {
        'name_sku': None,  # Номенклатура
        'units_of_measurement': None,  # Ед. изм.,
        'mark': None,  # метка
        'multiplicity': None,  # крат
        'expiration_dates': None,  # сроки
        'orders_is_on_the_way': {},  # заказ в пути
        'average_values': {},  # ср: {'10,11,': 28, ...}
        'comments': None,  # комментарии
    }

    columns_val = []
    columns_dict = []

    for row in range(1, max_row + 1):
        current_sku = None
        current_obj = None
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=row, column=col).value
            if not all(all_columns.values()):
                if cell_value == 'Номенклатура':
                    all_columns['name_sku'] = col
                    columns_val.append(col)
                if cell_value == 'Ед. изм.':
                    all_columns['units_of_measurement'] = col
                    columns_val.append(col)
                if cell_value == 'метка':
                    all_columns['mark'] = col
                    columns_val.append(col)
                if cell_value == 'крат':
                    all_columns['multiplicity'] = col
                    columns_val.append(col)
                if cell_value == 'сроки':
                    all_columns['expiration_dates'] = col
                    columns_val.append(col)
                if cell_value in ('заказ', 'заказ в пути'):
                    if ws.cell(row=row + 1, column=col).value:
                        all_columns['orders_is_on_the_way'].update({ws.cell(row=row + 1, column=col).value: col})
                        columns_dict.append(col)
                if cell_value in ('ср', 'ср нов'):
                    if ws.cell(row=row + 1, column=col).value:
                        all_columns['average_values'].update({ws.cell(row=row + 1, column=col).value: col})
                        columns_dict.append(col)
                if cell_value in ('комментарии', 'коментарии', 'комментарий'):
                    all_columns['comments'] = col
                    columns_val.append(col)
            if all(all_columns.values()) and col == all_columns['name_sku'] and cell_value:
                current_sku = cell_value
                if not current_sku in table.rows:  # Добавить СКЮ, которого нет в новом движении
                    obj = Row(current_sku)
                    obj.COLOR = RED_CELL
                    table.add_row(obj)
                else:
                    obj = table.rows[table.rows.index(current_sku)]
                    obj.COLOR = None
                current_obj = table.rows[table.rows.index(current_sku)]
            if all(all_columns.values()) and current_sku and col in columns_val:
                name_col = list(filter(lambda x: x[1] == col, all_columns.items()))[0][0]
                if name_col == 'name_sku':
                    continue
                current_obj.__dict__[name_col] = cell_value
            elif all(all_columns.values()) and current_sku and col in columns_dict:
                if col in all_columns['orders_is_on_the_way'].values():
                    value = cell_value
                    if value == '#REF!':
                        value = '#Н/Д'# 0
                    key_or_in_way = list(filter(lambda x: x[1] == col, all_columns['orders_is_on_the_way'].items()))[0][0]
                    current_obj.__dict__['orders_is_on_the_way'][key_or_in_way] = value
                    table.COLUMNS['orders_is_on_the_way'][1].update({key_or_in_way: None})
                elif col in all_columns['average_values'].values():
                    value = ws.cell(row=row, column=col).value
                    key_av_values = list(filter(lambda x: x[1] == col, all_columns['average_values'].items()))[0][0]
                    current_obj.__dict__['average_values'][key_av_values] = value
                    table.COLUMNS['average_values'][1].update({key_av_values: None})
                else:
                    raise Exception('NOT FOUND!!!')
    print(all_columns)


def add_data_from_sales_file(file, table, names):
    columns = {'Номенклатура': None}
    columns[names[0]] = None

    wb, ws, max_col, max_row = read_xlsx_file(file)
    for row in range(1, max_row + 1):
        current_obj = None
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value in columns:  # находим колонки: Номенклатура и количество
                if not columns[cell_value]:
                    columns[cell_value] = col  # связываем название колонки с ее номером
            if col in columns.values():
                if cell_value in table.rows:
                    current_obj = table.rows[table.rows.index(cell_value)]
                elif current_obj and cell_value:
                    current_obj.__dict__[names[1]] = cell_value


def create_new_file(name_file, table):
    last_row = 500
    width_first_col = 60
    width_else_cols = 8
    num_format = NUM_FORMAT
    all_columns = [j + chr(i) for j in ('', 'A') for i in range(65, 65 + 26)]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.zoomScale = 85 # масштаб страницы
    wb.add_named_style(new_style)
    for col in all_columns:
        ws[f'{col}{last_row}']
        for cell in ws[col]:
            cell.style = style_name
            cell.number_format = num_format
    for i in all_columns:
        ws.column_dimensions[i].width = width_else_cols
    ws.column_dimensions['A'].width = width_first_col
    create_header(table, ws)
    fill_file(table, ws, all_columns)
    wb.save(name_file)


def create_header(table, ws):
    header_color = HEADER_COLOR
    header_col_sum = HEADER_COLOR_SUM
    current_date = input('Введите текущую дату: ')
    table.number_columns()  # пронумеровать колонки
    header_data = table.get_header()
    for i in header_data:
        if type(header_data[i][1]) is dict:
            for j in header_data[i][1]:
                cell = ws.cell(row=table.START_HEADER, column=header_data[i][1][j])
                cell.value = header_data[i][0]
                cell.fill = PatternFill(patternType='solid', fgColor=header_color)
                cell.font = Font(bold=True)
                ws.cell(row=table.START_HEADER + 1, column=header_data[i][1][j]).value = j
                if header_data[i][2]: # добавление суммы в колонку
                    cell_sum = ws.cell(row=table.START_HEADER + 2, column=header_data[i][1][j])
                    sym = get_column_letter(header_data[i][1][j])
                    cell_sum.value = f'=SUM({sym}{table.START_ROWS}:{sym}{table.END_ROWS})'
                    cell_sum.fill = PatternFill(patternType='solid', fgColor=header_col_sum)
        else:
            cell = ws.cell(row=table.START_HEADER, column=header_data[i][1])
            cell.value = header_data[i][0]
            cell.fill = PatternFill(patternType='solid', fgColor=header_color)
            cell.font = Font(bold=True)
            if i == 'order':
                cell.font = Font(bold=True, color=RED_FONT) # меняем цвет шрифта
            if i == 'new_average_sales':
                ws.cell(row=table.START_HEADER + 1, column=header_data[i][1]).value = current_date
            if header_data[i][2]: # добавление суммы в колонку
                cell_sum = ws.cell(row=table.START_HEADER + 2, column=header_data[i][1])
                sym = get_column_letter(header_data[i][1])
                cell_sum.value = f'=SUM({sym}{table.START_ROWS}:{sym}{table.END_ROWS})'
                cell_sum.fill = PatternFill(patternType='solid', fgColor=header_col_sum)


def fill_file(table, ws, a_c):
    thick = Side(border_style="thin", color="000000")
    rows = table.get_rows()
    for row, obj_row in enumerate(rows, table.START_ROWS):
        for name_col in obj_row.get_columns():
            data = obj_row.__dict__[name_col]
            if not data is None:
                if type(data) is dict:
                    for key in data:
                        ws.cell(row=row, column=table.COLUMNS[name_col][1][key]).value = data[key]
                else:
                    if name_col == 'COLOR':
                        continue
                    cell = ws.cell(row=row, column=table.COLUMNS[name_col][1])
                    cell.value = data
                    if name_col == 'name' and obj_row.COLOR:
                        color = obj_row.COLOR
                        cell.fill = PatternFill(patternType='solid', fgColor=color)
        sale = a_c[table.COLUMNS['sales'][1] - 1]
        decl = a_c[table.COLUMNS['declared'][1] - 1]
        diff = a_c[table.COLUMNS['diff'][1] - 1]
        ws[f'{diff}{row}'] = f'={sale}{row}-{decl}{row}'
        cell_border1 = a_c[table.COLUMNS['order'][1] - 1]
        ws[f'{cell_border1}{row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
        cell_border2 = a_c[table.COLUMNS['order_from_f'][1] - 1]
        ws[f'{cell_border2}{row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
