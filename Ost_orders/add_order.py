import openpyxl


def read_file(file_name):
    workbook = openpyxl.load_workbook(file_name, data_only=True)

    sheet = workbook.active  # Получение активного листа

    rows = sheet.max_row
    columns = sheet.max_column

    head = 3

    orders = {}
    col_num = []

    substring = 'заказ'

    for cell in sheet[head]:
        if cell.value == substring:
            orders[cell.col_idx] = [sheet[f'{cell.column_letter}{head + 1}'].value, {}]
            col_num.append(cell.col_idx)

    for row in sheet[head + 1:rows + 1]:
        if row[0].value:
            for ind in col_num:
                if sheet.cell(row = row[0].row, column = ind).value:
                    orders[ind][1][sheet.cell(row = row[0].row, column = 1).value] = sheet.cell(row = row[0].row, column = ind).value
    return orders


def write_to_file(file_name, orders):
    workbook = openpyxl.load_workbook(file_name, data_only=True)

    sheet = workbook.active  # Получение активного листа

    rows = sheet.max_row
    columns = sheet.max_column

    for order in orders:
        # 19: ['07,12,', {'4558 ДОКТОРСКАЯ ГОСТ вар п/о  Останкино': 70, '5495 ВЕТЧ.С ИНДЕЙКОЙ Папа может п/о 400*6  Останкино': 120, ...
        col_add = 3 # C
        sheet.insert_cols(col_add)
        sheet.cell(row = 1, column = col_add).value = orders[order][0]
        sheet.cell(row = 2, column = col_add).value = f'=СУММ(C{3}:C{rows})'
        for row in sheet:
            if row[0].value in orders[order][1]:
                sheet.cell(row = row[0].row, column = col_add).value = orders[order][1][row[0].value]
                del orders[order][1][row[0].value]
        if orders[order][1]:
            create_error_file(orders[order])
    workbook.save(file_name)


def create_error_file(errors):
    with open(f'ERROR_{errors[0]}.txt', 'w') as f:
        f.write(str(errors[1]))


orders_file_name = input('Файл с заказами: ')
report_file_name = input('Файл с отчетом: ')

orders = read_file(orders_file_name)
write_to_file(report_file_name, orders)






