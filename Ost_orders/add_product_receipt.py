import openpyxl


def read_file(file_name):
    workbook = openpyxl.load_workbook(file_name, data_only=True)

    sheet = workbook.active  # Получение активного листа

    rows = sheet.max_row
    columns = sheet.max_column

    orders = {}
    col_num = []

    ignore = ['Поддон деревянный ЕВРО', 'Номенклатура', 'Европоддон (невозвратный)']

    for row in sheet[2:rows + 1]:
        product_name = row[0].value
        if product_name and not row[0].value in ignore:
            if not product_name in orders:
                orders[product_name] = float(row[1].value)
            else:
                orders[product_name] += float(row[1].value)
    return orders


def write_to_file(file_name, orders):
    workbook = openpyxl.load_workbook(file_name, data_only=True)

    get_date = input('Введите дату поступления товара: ')

    sheet = workbook.active  # Получение активного листа

    rows = sheet.max_row
    columns = sheet.max_column

    # {'6802 ОСТАНКИНСКАЯ вар п/о  Останкино': 17.094, 
    col_add = 3 # C
    sheet.insert_cols(col_add)
    sheet.cell(row = 1, column = col_add).value = f'{get_date},пр'
    sheet.cell(row = 2, column = col_add).value = f'=СУММ(C{3}:C{rows})'
    for row in sheet:
        if row[0].value in orders:
            sheet.cell(row = row[0].row, column = col_add).value = round(orders[row[0].value])
            del orders[row[0].value]
    if orders:
        create_error_file(orders, get_date)
    workbook.save(file_name)


def create_error_file(errors, get_date):
    with open(f'ERROR_{get_date}.txt', 'w') as f:
        f.write(str(errors))


orders_file_name = input('Файл с поступлением товара: ')
report_file_name = input('Файл с отчетом: ')

orders = read_file(orders_file_name)
write_to_file(report_file_name, orders)






