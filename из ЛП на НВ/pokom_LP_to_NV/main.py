import openpyxl


ERROR = []
COLUMN = 'X'

report = []
total_weight = 0
transferred_weight = 0
sku_svarog = {}
file_svarog = 'data/LP_to_NV.xlsx'

wb_svarog = openpyxl.load_workbook(file_svarog)
ws_svarog = wb_svarog.active
svarog = ws_svarog.iter_rows(1, ws_svarog.max_row)
for num, i in enumerate(svarog, 1):
    sku_svarog[i[0].value] = [0, False, i[1].value, None] # кол-во в заказе; найдено в файле Крым; название СКЮ; № строки в файле Крыма

file_Sochi = input('Введите названия файла для Логистического Партнера (отправитель): ')
file_Krim = input('Введите названия файла для Нового Времени (получатель): ')

wb_Krim = openpyxl.load_workbook(file_Krim)
ws_Krim = wb_Krim.active
Krim = ws_Krim.iter_rows(1, ws_Krim.max_row)
for num, i in enumerate(Krim, 1):
    if i[0].value in sku_svarog:
        sku_svarog[i[0].value][1] = True
        sku_svarog[i[0].value][3] = num

print(sku_svarog)

wb_Sochi = openpyxl.load_workbook(file_Sochi)
ws_Sochi = wb_Sochi.active
Sochi = ws_Sochi.iter_rows(1, ws_Sochi.max_row)
for num, i in enumerate(Sochi, 1):
    adress_Sochi = f'{COLUMN}{num}'
    if i[0].value in sku_svarog and ws_Sochi[adress_Sochi].value: # найден штрих-код и заказ не 0
        print(f'FIND => {i[0].value = }')
        print(f'{total_weight = } # {ws_Sochi[adress_Sochi].value = } # {adress_Sochi}')
        total_weight += ws_Sochi[adress_Sochi].value
        if sku_svarog[i[0].value][1]: # товар есть в крымском файле
            adress_Krim = f'{COLUMN}{sku_svarog[i[0].value][3]}'
            ws_Krim[adress_Krim] = ws_Sochi[adress_Sochi].value
            transferred_weight += ws_Sochi[adress_Sochi].value
            ws_Sochi[adress_Sochi] = 0
            report.append(f'{sku_svarog[i[0].value][2]} со штрих-кодом {i[0].value} перенесено из ячейки {adress_Sochi} в ячейку {adress_Krim}')
        else:
            ERROR.append(f'{sku_svarog[i[0].value][2]} со штрих-кодом {i[0].value} в бланке {file_Krim} не найдена!')

wb_Krim.save(file_Krim)
wb_Sochi.save(file_Sochi)

with open('REPORT.txt', 'w') as f:
    if ERROR:
        f.write(f'Обнаружены следующие ошибки:\n')
        for i in ERROR:
            f.write(i + '\n')
        f.write('#' * 50 + '\n\n')
    for i in report:
        f.write(i + '\n')
    f.write('\n')
    f.write(f'Из {total_weight}кг перенесено в бланк Крыма {transferred_weight}кг.')

